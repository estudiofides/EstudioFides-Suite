"""
Ficha de cliente: un Excel con diseño de formulario, uno por cliente,
guardado DENTRO de la carpeta del cliente ("Ficha_<Cliente>.xlsx").

Es la fuente principal de los datos de contacto de cada cliente
(apellido y nombre, dirección, teléfono, email, ciudad, notas) y de la
lista de sus expedientes (nombre + CUIJ + fecha de alta) -- eso no
vivía en ningún lado antes. La base SQLite del programa (src/database.py,
organizador.db) sigue siendo solo el caché de archivos ya analizados
(qué archivo es de qué cliente, qué número de expediente ya se vio):
no guarda datos de contacto, así que no hay dos lugares peleando por
ser "la verdad" de lo mismo.

Layout fijo (siempre las mismas filas/columnas, para poder leerlo de
vuelta sin ambigüedad):

    Fila 1: título "FICHA DE CLIENTE"
    Fila 2: subtítulo "Estudio Fides"
    Fila 3: (vacía)
    Filas 4-9: una por cada campo de CAMPOS, etiqueta en A, valor en B
    Fila 10: (vacía)
    Fila 11: encabezado de sección "EXPEDIENTES"
    Fila 12: encabezados de la tabla (Nombre del expediente / CUIJ / Fecha de alta)
    Fila 13 en adelante: un expediente por fila, hasta la primera fila
        vacía en la columna A

Vive en organizador_clientes/src/ para poder importarse desde
cualquier pestaña (pestana_alta_cliente.py, y a futuro cualquiera que
necesite mostrar/autocompletar datos de contacto de un cliente).
"""
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

CAMPOS = ["Apellido y Nombre", "Dirección", "Teléfono", "Email", "Ciudad", "Notas", "DNI/CUIT"]

# Campos que alcanzan como mínimo para considerar la ficha "completa"
# (además de Apellido y Nombre, que siempre es obligatorio): con al
# menos uno de estos, el cliente es contactable. Se usa en
# estado_ficha() -- ver Panel de Estado.
CAMPOS_CONTACTO = ["Dirección", "Teléfono", "Email"]

# IMPORTANTE: los campos nuevos SIEMPRE se agregan al FINAL de esta
# lista, nunca en el medio -- leer_ficha() lee cada campo por su
# posición (fila = _FILA_CAMPOS_DESDE + índice), así que insertar uno
# en el medio correría todos los que vienen después y desalinearía la
# lectura de fichas ya guardadas con el orden viejo.

_FILA_CAMPOS_DESDE = 4
_FILA_SECCION_EXPEDIENTES = _FILA_CAMPOS_DESDE + len(CAMPOS) + 1   # 11
_FILA_ENCABEZADO_EXPEDIENTES = _FILA_SECCION_EXPEDIENTES + 1        # 12
_FILA_EXPEDIENTES_DESDE = _FILA_ENCABEZADO_EXPEDIENTES + 1          # 13

_FILL_TITULO = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_FILL_SECCION = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_FUENTE_TITULO = Font(bold=True, size=14, color="FFFFFF")
_FUENTE_SECCION = Font(bold=True, size=11)
_FUENTE_ETIQUETA = Font(bold=True)


def ruta_ficha(carpeta_cliente):
    """Ruta donde vive (o viviría) la ficha de este cliente."""
    return carpeta_cliente / f"Ficha_{carpeta_cliente.name}.xlsx"


def _fila_seccion_expedientes(ws):
    """Busca en QUÉ fila está realmente el encabezado "EXPEDIENTES" en
    esta hoja puntual, en vez de asumir un número fijo -- así, si
    CAMPOS crece en el futuro, una ficha vieja (guardada con menos
    campos) sigue leyendo bien sus expedientes, en vez de desalinearse.
    Devuelve None si no se encuentra (archivo corrupto/inesperado)."""
    fila = _FILA_CAMPOS_DESDE
    limite = _FILA_CAMPOS_DESDE + len(CAMPOS) + 50  # margen de sobra
    while fila < limite:
        if ws.cell(row=fila, column=1).value == "EXPEDIENTES":
            return fila
        fila += 1
    return None


def leer_ficha(carpeta_cliente):
    """
    Devuelve {"datos": {campo: valor, ...}, "expedientes": [{"nombre",
    "cuij", "fecha_alta"}, ...]}, o None si el cliente todavía no
    tiene ficha (por ejemplo, uno viejo creado antes de esta pestaña).
    """

    ruta = ruta_ficha(carpeta_cliente)
    if not ruta.is_file():
        return None

    try:
        wb = load_workbook(ruta)
        ws = wb["Ficha"]
    except Exception:
        return None

    datos = {}
    for i, campo in enumerate(CAMPOS):
        valor = ws.cell(row=_FILA_CAMPOS_DESDE + i, column=2).value
        datos[campo] = valor if valor is not None else ""

    expedientes = []
    fila_seccion = _fila_seccion_expedientes(ws)
    if fila_seccion is not None:
        fila = fila_seccion + 2  # +1 encabezados de tabla, +1 primera fila de datos
        while True:
            nombre = ws.cell(row=fila, column=1).value
            if not nombre:
                break
            expedientes.append({
                "nombre": nombre,
                "cuij": ws.cell(row=fila, column=2).value or "",
                "fecha_alta": ws.cell(row=fila, column=3).value or "",
            })
            fila += 1

    return {"datos": datos, "expedientes": expedientes}


def estado_ficha(carpeta_cliente):
    """
    Devuelve "sin_ficha", "incompleta", o "completa":

      - "sin_ficha": el cliente todavía no tiene ninguna ficha creada.
      - "incompleta": tiene ficha, pero le falta el nombre, o no tiene
        NINGÚN dato de contacto (ni dirección, ni teléfono, ni email).
      - "completa": tiene nombre y al menos un dato de contacto.

    Se usa en Panel de Estado para el resumen de "cuántos clientes
    tienen la ficha completa" (ver README).
    """
    ficha = leer_ficha(carpeta_cliente)
    if not ficha:
        return "sin_ficha"

    datos = ficha["datos"]

    if not (datos.get("Apellido y Nombre") or "").strip():
        return "incompleta"

    if not any((datos.get(campo) or "").strip() for campo in CAMPOS_CONTACTO):
        return "incompleta"

    return "completa"


def guardar_ficha(carpeta_cliente, datos, expedientes_nuevos=None):
    """
    Crea (si no existía) o reescribe la ficha del cliente, dentro de
    su propia carpeta.

    `datos`: dict con algunas (o todas) las claves de CAMPOS -- las
    que falten quedan vacías.

    `expedientes_nuevos`: lista opcional de {"nombre", "cuij"} para
    AGREGAR a los que ya hubiera en la ficha (no duplica si ya existía
    exactamente el mismo nombre+CUIJ). Los expedientes que ya estaban
    en la ficha (si la había) se conservan siempre.

    Devuelve la ruta del archivo guardado.
    """

    existente = leer_ficha(carpeta_cliente)
    expedientes = list(existente["expedientes"]) if existente else []

    if expedientes_nuevos:
        ya = {(e["nombre"], e.get("cuij", "")) for e in expedientes}
        for nuevo in expedientes_nuevos:
            clave = (nuevo.get("nombre", ""), nuevo.get("cuij", ""))
            if not clave[0] or clave in ya:
                continue
            expedientes.append({
                "nombre": nuevo.get("nombre", ""),
                "cuij": nuevo.get("cuij", ""),
                "fecha_alta": nuevo.get("fecha_alta") or date.today().strftime("%d/%m/%Y"),
            })
            ya.add(clave)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ficha"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18

    ws.merge_cells("A1:C1")
    ws["A1"] = "FICHA DE CLIENTE"
    ws["A1"].font = _FUENTE_TITULO
    ws["A1"].fill = _FILL_TITULO
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    ws["A2"] = "Estudio Fides"
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    for i, campo in enumerate(CAMPOS):
        fila = _FILA_CAMPOS_DESDE + i
        celda_etiqueta = ws.cell(row=fila, column=1, value=f"{campo}:")
        celda_etiqueta.font = _FUENTE_ETIQUETA
        celda_valor = ws.cell(row=fila, column=2, value=(datos.get(campo) or ""))
        if campo == "Notas":
            celda_valor.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[fila].height = 45

    ws.merge_cells(
        start_row=_FILA_SECCION_EXPEDIENTES, start_column=1,
        end_row=_FILA_SECCION_EXPEDIENTES, end_column=3,
    )
    celda = ws.cell(row=_FILA_SECCION_EXPEDIENTES, column=1, value="EXPEDIENTES")
    celda.font = _FUENTE_SECCION
    celda.fill = _FILL_SECCION
    celda.alignment = Alignment(horizontal="center")

    for col, texto in zip("ABC", ["Nombre del expediente", "CUIJ", "Fecha de alta"]):
        c = ws[f"{col}{_FILA_ENCABEZADO_EXPEDIENTES}"]
        c.value = texto
        c.font = _FUENTE_ETIQUETA

    fila = _FILA_EXPEDIENTES_DESDE
    for exp in expedientes:
        ws.cell(row=fila, column=1, value=exp["nombre"])
        ws.cell(row=fila, column=2, value=exp["cuij"])
        ws.cell(row=fila, column=3, value=exp["fecha_alta"])
        fila += 1

    ruta = ruta_ficha(carpeta_cliente)
    wb.save(ruta)
    return ruta
