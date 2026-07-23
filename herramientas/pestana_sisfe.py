"""
Pestaña "SISFE" para la app unificada (hub_app.py).
Misma lógica de herramientas/extraer_datos_sisfe.py, embebida como Frame.
"""
import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import tkinter.scrolledtext as scrolledtext

from pypdf import PdfReader
from openpyxl import Workbook, load_workbook

from src.tooltip import agregar_tooltip
from src.config import CIUDADES


COLUMNAS = ["Carátula", "Expte/CUIJ", "Juzgado", "Inicio", "Ultima_Act"]


def extraer_datos_pdf(ruta_pdf):

    expedientes = []

    with open(ruta_pdf, 'rb') as f:
        lector = PdfReader(f)
        texto = ""
        for pag in lector.pages:
            t = pag.extract_text()
            if t:
                texto += t + "\n"

    lineas = texto.split('\n')
    datos = None

    for linea in lineas:
        linea = linea.strip()

        if linea.startswith("Carátula:"):
            datos = {
                "Carátula": linea.replace("Carátula:", "").strip(),
                "Expte/CUIJ": "",
                "Juzgado": "",
                "Inicio": "",
                "Ultima_Act": ""
            }

        elif datos and "Radicación Actual:" in linea:
            datos["Juzgado"] = linea.replace("Radicación Actual:", "").strip()

        elif datos and "Expte:" in linea:
            cuij_match = re.search(r'(\d{2}-\d{8}-\d)', linea)
            datos["Expte/CUIJ"] = cuij_match.group(1) if cuij_match else ""

        elif datos and "Fecha de Inicio:" in linea:
            fechas = re.findall(r'(\d{2}/\d{2}/\d{4})', linea)
            if len(fechas) >= 1:
                datos["Inicio"] = fechas[0]
            if len(fechas) >= 2:
                datos["Ultima_Act"] = fechas[1]

            expedientes.append(datos)
            datos = None

    return expedientes


def _nombre_hoja(ciudad):
    """Nombre de ciudad "pelado" (sin el "N.- " adelante) para usar
    como nombre de hoja -- más corto y legible que el nombre completo,
    y de todas formas evita choques porque cada ciudad es única."""
    return ciudad.split(".- ", 1)[-1].strip() if ".- " in ciudad else ciudad.strip()


def _obtener_o_crear_hoja(wb, ciudad):
    """Cada ciudad tiene su propia hoja dentro del mismo Excel (para
    tener TODOS los expedientes en un solo archivo, uno por ciudad,
    como pidió Leandro). La crea con encabezados si todavía no existe."""

    nombre_hoja = _nombre_hoja(ciudad)

    if nombre_hoja in wb.sheetnames:
        return wb[nombre_hoja]

    ws = wb.create_sheet(title=nombre_hoja)
    ws.append(COLUMNAS)

    # Si el libro se acaba de crear, openpyxl le puso una hoja "Sheet"
    # vacía por defecto -- si no se usó para nada, se saca para no
    # dejar una hoja fantasma sin datos.
    if "Sheet" in wb.sheetnames and wb["Sheet"] is not ws:
        vacia = wb["Sheet"]
        if vacia.max_row <= 1 and vacia.max_column <= 1 and vacia["A1"].value is None:
            del wb["Sheet"]

    return ws


def guardar_en_excel(expedientes, ruta_excel, ciudad):

    if os.path.exists(ruta_excel):
        wb = load_workbook(ruta_excel)
    else:
        wb = Workbook()

    ws = _obtener_o_crear_hoja(wb, ciudad)

    existentes = {}
    for fila_num in range(2, ws.max_row + 1):
        cuij = ws.cell(row=fila_num, column=2).value
        if cuij:
            existentes[cuij] = fila_num

    nuevos = 0
    actualizados = 0

    for exp in expedientes:

        cuij = exp.get("Expte/CUIJ", "")
        fila_datos = [exp.get(c, "") for c in COLUMNAS]

        if cuij and cuij in existentes:
            fila_num = existentes[cuij]
            for col, valor in enumerate(fila_datos, start=1):
                ws.cell(row=fila_num, column=col, value=valor)
            actualizados += 1
        else:
            ws.append(fila_datos)
            if cuij:
                existentes[cuij] = ws.max_row
            nuevos += 1

    ws.auto_filter.ref = ws.dimensions
    wb.save(ruta_excel)

    return nuevos, actualizados


def construir_pestana(parent):

    frame = ttk.Frame(parent, padding=15)

    ttk.Label(frame, text="📊 Extractor de Expedientes SISFE", font=("Arial", 14, "bold")).pack(pady=(5, 5))
    ttk.Label(
        frame,
        text="Extrae Carátula, Expte/CUIJ, Juzgado y fechas de un PDF, y los guarda en un Excel -- "
             "un solo archivo para todas las ciudades, con una hoja por ciudad.",
        foreground="#555555"
    ).pack(pady=(0, 15))

    frame_excel = ttk.Frame(frame)
    frame_excel.pack(fill=tk.X, pady=(0, 10))

    ttk.Label(frame_excel, text="Base Excel:").pack(side=tk.LEFT)
    entry_excel = tk.Entry(frame_excel, width=48)
    entry_excel.insert(0, os.path.join(os.getcwd(), "Base_Expedientes_SISFE.xlsx"))
    entry_excel.pack(side=tk.LEFT, padx=8)

    def elegir_excel():
        ruta = filedialog.asksaveasfilename(
            title="Elegí o creá el archivo Excel de expedientes",
            initialfile="Base_Expedientes_SISFE.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if ruta:
            entry_excel.delete(0, tk.END)
            entry_excel.insert(0, ruta)

    btn_elegir_excel = ttk.Button(frame_excel, text="Elegir...", command=elegir_excel)
    btn_elegir_excel.pack(side=tk.LEFT)
    agregar_tooltip(btn_elegir_excel, "Elegí dónde está (o dónde se va a crear) el Excel donde se "
                                        "van guardando los expedientes extraídos (TODAS las ciudades "
                                        "en el mismo archivo, cada una en su hoja).")

    frame_ciudad = ttk.Frame(frame)
    frame_ciudad.pack(fill=tk.X, pady=(0, 15))

    ttk.Label(frame_ciudad, text="Ciudad del PDF:").pack(side=tk.LEFT)
    combo_ciudad = ttk.Combobox(frame_ciudad, values=CIUDADES, width=25, state="readonly")
    combo_ciudad.pack(side=tk.LEFT, padx=8)
    agregar_tooltip(combo_ciudad, "A qué ciudad corresponden los expedientes de este PDF -- se "
                                    "guardan en la hoja de esa ciudad dentro del Excel.")

    btn_iniciar = ttk.Button(frame, text="📄 Seleccionar PDF de SISFE y Procesar")
    btn_iniciar.pack(pady=8)
    agregar_tooltip(btn_iniciar, "Elegí el PDF descargado de SISFE: extrae Carátula, Expte/CUIJ, "
                                   "Juzgado y fechas, y los agrega (o actualiza) en la hoja de la "
                                   "ciudad elegida arriba.")

    consola = scrolledtext.ScrolledText(frame, wrap=tk.WORD, font=("Consolas", 10), bg="#f4f4f4")
    consola.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
    consola.insert(tk.END, "Listo. Elegí tu Excel de expedientes, la ciudad, y seleccioná un PDF.\n\n")

    def proceso(ruta_pdf, ruta_excel, ciudad):

        consola.insert(tk.END, f"► Leyendo: {os.path.basename(ruta_pdf)} (hoja: {_nombre_hoja(ciudad)})\n")
        consola.see(tk.END)

        try:
            expedientes = extraer_datos_pdf(ruta_pdf)
        except Exception as e:
            consola.insert(tk.END, f"\n❌ Error al leer el PDF: {e}\n")
            consola.insert(tk.END, "-" * 60 + "\n\n")
            btn_iniciar.config(state=tk.NORMAL)
            return

        if not expedientes:
            consola.insert(tk.END, "\n⚠️ No se encontró ningún expediente con el formato esperado.\n")
            consola.insert(tk.END, "   (Se buscan las etiquetas 'Carátula:', 'Radicación Actual:',\n")
            consola.insert(tk.END, "   'Expte:' y 'Fecha de Inicio:' en el texto del PDF.)\n")
            consola.insert(tk.END, "-" * 60 + "\n\n")
            btn_iniciar.config(state=tk.NORMAL)
            return

        consola.insert(tk.END, f"► Se encontraron {len(expedientes)} expediente(s). Guardando en Excel...\n")
        consola.see(tk.END)

        try:
            nuevos, actualizados = guardar_en_excel(expedientes, ruta_excel, ciudad)
        except Exception as e:
            consola.insert(tk.END, f"\n❌ Error al guardar en Excel: {e}\n")
            consola.insert(tk.END, "   (¿Está el Excel abierto en otra ventana? Cerralo e intentá de nuevo.)\n")
            consola.insert(tk.END, "-" * 60 + "\n\n")
            btn_iniciar.config(state=tk.NORMAL)
            return

        consola.insert(tk.END, f"\n✅ LISTO. Nuevos: {nuevos} | Actualizados: {actualizados}\n")
        consola.insert(tk.END, f"📄 Base de datos: {ruta_excel}\n")
        consola.insert(tk.END, "-" * 60 + "\n\n")
        consola.see(tk.END)

        btn_iniciar.config(state=tk.NORMAL)

    def iniciar():

        ruta_excel = entry_excel.get().strip()
        if not ruta_excel:
            messagebox.showwarning(
                "Falta la base",
                "Elegí primero dónde está (o se va a crear) el Excel de expedientes."
            )
            return

        ciudad = combo_ciudad.get().strip()
        if not ciudad:
            messagebox.showwarning("Falta la ciudad", "Elegí a qué ciudad corresponde este PDF.")
            return

        ruta_pdf = filedialog.askopenfilename(
            title="Seleccioná el PDF descargado de SISFE",
            filetypes=[("Archivos PDF", "*.pdf")]
        )
        if not ruta_pdf:
            return

        btn_iniciar.config(state=tk.DISABLED)
        consola.insert(tk.END, "\n" + "=" * 60 + "\n")

        hilo = threading.Thread(target=proceso, args=(ruta_pdf, ruta_excel, ciudad), daemon=True)
        hilo.start()

    btn_iniciar.config(command=iniciar)

    return frame
